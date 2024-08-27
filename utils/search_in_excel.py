import openpyxl
from utils.consts import KEYS

def find_user_code_range(file_path, user_code, column='G') -> str:
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    col_num = openpyxl.utils.column_index_from_string(column)

    for row in sheet.iter_rows(min_col=col_num, max_col=col_num, values_only=False):
        cell = row[0]
        if cell.value == user_code:
            row_values = [sheet.cell(row=cell.row, column=col).value for col in range(1, sheet.max_column + 1)]
            result = dict(zip(KEYS, row_values))
            format = data_formatting(result)
            return format

    return 'Уточните номер СНИЛС'

def data_formatting(dict) -> str:
    return (
        f"*Название группы:* {dict['study_group']} 🏫\n\n"
        f"*Направление обучения:* {dict['specialization']} 🎓\n\n"
        f"*Профиль обучения:* {dict['specialty']} 🖥️\n\n"
        f"*Куратор группы:* {dict['tutor']} 👨‍🏫\n\n"
        f"*Кафедра:* {dict['department']} 🏢\n\n"
        f"*Заведующий кафедры:* {dict['head_department']} 👩‍🏫\n\n"
        f"*Страница кафедры:* {dict['department_link']} 🌐\n\n"
        f"*Контактная информация:* {dict['contacts']} 📧\n\n"
        f"*Ссылка на расписание:* {dict['link_schedule']} 📅"
    )
